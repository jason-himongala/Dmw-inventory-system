"""Inspect PEOS Monitoring.xlsx and print a sample of its contents.

Usage:
  python scripts/inspect_peos.py

Outputs:
- sheets available in the workbook
- active sheet name
- first N rows of the active sheet (default 10)
"""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any

import openpyxl


def main() -> None:
    root = Path(__file__).resolve().parents[1]
    xlsx_path = root / "resources" / "exel" / "PEOS Monitoring.xlsx"

    if not xlsx_path.exists():
        raise SystemExit(f"Could not find file: {xlsx_path}")

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    print("sheets:", wb.sheetnames)
    print("active:", ws.title)

    rows = list(ws.iter_rows(values_only=True))
    sample = rows[:10]

    def _clean(value: Any) -> Any:
        if value is None:
            return None
        return value

    print("first 10 rows:")
    print(json.dumps([[ _clean(c) for c in row ] for row in sample], ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
